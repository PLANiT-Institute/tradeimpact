# SPDX-License-Identifier: GPL-3.0-or-later
"""Load the TI data workbook and reference DB into validated models.

Two entry points:
    load_workbook_inputs(path) -> WorkbookInputs  (the TI_Data_Workbook template)
    load_reference_db(path)    -> dict[code, Country]  (the seed reference values)

Both are tolerant of the template's empty / placeholder cells: missing inputs become
``None`` and are surfaced later, never silently defaulted.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from ti_framework.io import schema as S
from ti_framework.models import (
    Country,
    DataTier,
    ScenarioRate,
    SupportParams,
    Vehicle,
    Volume,
)

PRORATA_IDENTITY_WARNING = (
    "PRORATA_IDENTITY: r_fleet=r_power=economy-wide rate (sector-split factor 1.0). "
    "Derive independently to break the identity (NOTES.md D1). Benchmark tier downgraded."
)


@dataclass
class WorkbookInputs:
    """Everything read from the TI_Data_Workbook template."""

    countries: dict[str, Country] = field(default_factory=dict)
    vehicles: list[Vehicle] = field(default_factory=list)
    volumes: list[Volume] = field(default_factory=list)
    support: SupportParams = field(default_factory=SupportParams)
    missing_inputs: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    header = next(ws.iter_rows(values_only=True))
    out: dict[str, int] = {}
    for i, c in enumerate(header):
        t = S.text(c)
        if t:
            out[t] = i
    return out


def _parse_range(value: object) -> tuple[float | None, float | None]:
    """Parse a textual range like '0.53–0.61' or '4.34-5.39' into (low, high)."""
    t = S.text(value)
    if not t or S.is_placeholder(value):
        return (None, None)
    norm = t.replace("–", "-").replace("—", "-").replace("%", "").strip()
    parts = [p for p in norm.split("-") if p.strip()]
    try:
        nums = [float(p) for p in parts]
    except ValueError:
        return (None, None)
    if not nums:
        return (None, None)
    return (nums[0], nums[-1])


def _apply_ndc_defaults(country: Country) -> None:
    """Apply confirmed decisions D1/D2 to fill S2 rates from the economy-wide rate.

    D2: S2 central = unconditional (low) economy-wide rate; conditional (high) -> S2 upper
    sensitivity bound. D1: with no independent power rate, r_power S2 = r_fleet S2 (pro-rata
    identity) + a loud warning + tier downgrade.
    """
    if country.is_flag:
        return  # D3: never fabricate an S2 for FLAG markets
    low = country.econ_rate_low
    high = country.econ_rate_high
    # Only default where the explicit S2 cell was empty.
    if country.r_fleet.s2 is None and low is not None:
        country.r_fleet.s2 = low
        country.r_fleet.s2_upper = high
    if country.r_power.s2 is None and low is not None:
        country.r_power.s2 = low
        country.r_power.s2_upper = high
        # D1 identity warning (only meaningful when both came from the same econ-wide rate)
        if country.r_fleet.s2 == country.r_power.s2 and low is not None:
            country.warnings.append(PRORATA_IDENTITY_WARNING)
            if country.tier in (DataTier.A, DataTier.UNKNOWN):
                country.tier = DataTier.B  # downgrade


def _load_layer1(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Country]:
    hm = _header_map(ws)

    def col(name: str) -> int | None:
        return hm.get(name)

    def get(row: tuple, name: str) -> object:
        idx = col(name)
        return row[idx] if idx is not None and idx < len(row) else None

    out: dict[str, Country] = {}
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r == 0:
            continue
        code = S.text(get(row, "Code"))
        name = S.text(get(row, "Operating country"))
        if not code or not name:
            continue
        grid = S.gco2_per_kwh_to_kg(
            S.num(get(row, "Grid intensity G_c(0) gCO2/kWh (2024)"))
        )
        country = Country(
            name=name,
            code=code,
            grid_intensity=grid,
            ndc_headline=S.text(get(row, "NDC headline target")),
            base_year=int(S.num(get(row, "Base yr")) or 0) or None,
            target_year=int(S.num(get(row, "Target yr")) or 0) or None,
            reduction_low=S.num(get(row, "Reduction low (frac)")),
            reduction_high=S.num(get(row, "Reduction high (frac)")),
            econ_rate_low=S.parse_rate_fraction(get(row, "Econ-wide rate low (%/yr)")),
            econ_rate_high=S.parse_rate_fraction(get(row, "Econ-wide rate high (%/yr)")),
            r_fleet=ScenarioRate(
                s1=S.parse_rate_fraction(get(row, "r_fleet S1 (STEPS)")),
                s2=S.parse_rate_fraction(get(row, "r_fleet S2 (NDC)")),
                s3=S.parse_rate_fraction(get(row, "r_fleet S3 (NZE)")),
            ),
            r_power=ScenarioRate(
                s1=S.parse_rate_fraction(get(row, "r_power S1 (STEPS)")),
                s2=S.parse_rate_fraction(get(row, "r_power S2 (NDC)")),
                s3=S.parse_rate_fraction(get(row, "r_power S3 (NZE)")),
            ),
            status=S.parse_status(get(row, "Benchmark status")),
            source=S.text(get(row, "Source")),
            tier=DataTier.A,  # reference grid values are Tier A; refined below
        )
        _apply_ndc_defaults(country)
        out[code] = country
    return out


def _load_layer2(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[Vehicle]:
    hm = _header_map(ws)

    def get(row: tuple, name: str) -> object:
        idx = hm.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    out: list[Vehicle] = []
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r == 0:
            continue
        brand = S.text(get(row, "Brand"))
        pt = S.parse_powertrain(get(row, "Powertrain (BEV/PHEV/HEV/ICE)"))
        if not brand and pt is None:
            continue  # wholly empty template row
        out.append(
            Vehicle(
                brand=brand or "",
                model=S.text(get(row, "Model")),
                powertrain=pt,  # type: ignore[arg-type]
                segment=S.text(get(row, "Segment")),
                eta_ev=S.num(get(row, "EV efficiency (kWh/km)")),
                ice_intensity=S.gco2_per_km_to_kg(
                    S.num(get(row, "ICE emission intensity (gCO2/km)"))
                ),
                uf=S.num(get(row, "PHEV utility factor")),
                realworld_correction=S.num(get(row, "Real-world correction factor")),
                tier=S.parse_tier(get(row, "Tier (A/B/C)")),
                source=S.text(get(row, "Source")),
            )
        )
    return out


def _load_volumes(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[Volume]:
    hm = _header_map(ws)

    def get(row: tuple, name: str) -> object:
        idx = hm.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    out: list[Volume] = []
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r == 0:
            continue
        code = S.text(get(row, "Code"))
        units = S.num(get(row, "Units"))
        pt = S.parse_powertrain(get(row, "Powertrain"))
        if not code:
            continue
        out.append(
            Volume(
                country_code=code,
                year=int(S.num(get(row, "Year")) or 0),
                brand=S.text(get(row, "Brand")),
                model=S.text(get(row, "Model")),
                powertrain=pt,  # type: ignore[arg-type]
                units=units,
                segment=S.text(get(row, "Segment")),
                source=S.text(get(row, "Source DB")),
                tier=S.parse_tier(get(row, "Tier")),
            )
        )
    return out


def _load_support(ws: openpyxl.worksheet.worksheet.Worksheet | None) -> SupportParams:
    sp = SupportParams()
    if ws is None:
        return sp
    hm = _header_map(ws)

    def get(row: tuple, name: str) -> object:
        idx = hm.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r == 0:
            continue
        sym = S.text(get(row, "Symbol"))
        sector = S.text(get(row, "Sector"))
        central = S.num(get(row, "Central value"))
        if sym == "T" and sector == "Automotive" and central is not None:
            sp.lifetime_T = int(central)
    return sp


def load_workbook_inputs(path: str | Path) -> WorkbookInputs:
    """Load the TI_Data_Workbook template into validated models, tolerant of empty cells."""
    wb = openpyxl.load_workbook(path, data_only=True)
    inputs = WorkbookInputs()

    if S.SHEET_LAYER1 in wb.sheetnames:
        inputs.countries = _load_layer1(wb[S.SHEET_LAYER1])
    if S.SHEET_LAYER2 in wb.sheetnames:
        inputs.vehicles = _load_layer2(wb[S.SHEET_LAYER2])
    if S.SHEET_REG in wb.sheetnames:
        inputs.volumes = _load_volumes(wb[S.SHEET_REG])
    inputs.support = _load_support(wb[S.SHEET_SUPPORT] if S.SHEET_SUPPORT in wb.sheetnames else None)

    # Collect warnings and missing-input markers.
    for c in inputs.countries.values():
        inputs.warnings.extend(f"[{c.code}] {w}" for w in c.warnings)
        if c.is_flag:
            inputs.missing_inputs.append(f"[{c.code}] benchmark FLAG: {c.status.value}")
    if not any(v.units for v in inputs.volumes):
        inputs.missing_inputs.append("Registration_Vcv: no volume units collected")
    if not any(v.eta_ev or v.ice_intensity for v in inputs.vehicles):
        inputs.missing_inputs.append("Layer2_vehicle_params: no vehicle parameters collected")
    if inputs.support.lifetime_T is None:
        inputs.missing_inputs.append("Support_params: vehicle lifetime T not collected")
    return inputs


def load_reference_db(path: str | Path) -> dict[str, Country]:
    """Load the seed reference DB (D3 grid + D2 NDC targets) into Country objects."""
    wb = openpyxl.load_workbook(path, data_only=True)
    countries: dict[str, Country] = {}

    if "D3_grid_intensity" in wb.sheetnames:
        ws = wb["D3_grid_intensity"]
        hm = _header_map(ws)

        def g3(row: tuple, name: str) -> object:
            idx = hm.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r == 0:
                continue
            code = S.text(g3(row, "Code"))
            name = S.text(g3(row, "Economy"))
            if not code or not name:
                continue
            grid = S.gco2_per_kwh_to_kg(
                S.num(g3(row, "Grid carbon intensity G_c(0) (gCO2/kWh)"))
            )
            countries[code] = Country(
                name=name,
                code=code,
                grid_intensity=grid,
                tier=S.parse_tier(g3(row, "Tier")),
                source=S.text(g3(row, "Source")),
            )

    if "D2_NDC_targets" in wb.sheetnames:
        ws = wb["D2_NDC_targets"]
        hm = _header_map(ws)

        def g2(row: tuple, name: str) -> object:
            idx = hm.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r == 0:
                continue
            code = S.text(g2(row, "Code"))
            if not code:
                continue
            c = countries.setdefault(code, Country(name=S.text(g2(row, "Economy")) or code, code=code))
            c.ndc_headline = S.text(g2(row, "Latest NDC headline target"))
            c.base_year = int(S.num(g2(row, "Base year")) or 0) or None
            c.target_year = int(S.num(g2(row, "Target year")) or 0) or None
            red_lo, red_hi = _parse_range(g2(row, "Reduction (frac, low–high)"))
            c.reduction_low, c.reduction_high = red_lo, red_hi
            er_lo, er_hi = _parse_range(g2(row, "Economy-wide r, S2 (%/yr)"))
            c.econ_rate_low = None if er_lo is None else er_lo / 100.0
            c.econ_rate_high = None if er_hi is None else er_hi / 100.0
            c.status = S.parse_status(g2(row, "Status"))
            c.source = S.text(g2(row, "Source"))
            _apply_ndc_defaults(c)

    return countries
