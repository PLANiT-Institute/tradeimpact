# SPDX-License-Identifier: GPL-3.0-or-later
"""CSV adapter: one-CSV-per-sheet loads identically to the xlsx, against one schema.

Also holds the no-hardcode guarantee (build brief §3): changing a value in the input
data changes the engine output with zero code edits.
"""
from pathlib import Path

import openpyxl
import pytest

from ti_framework.core.aggregate import Placement
from ti_framework.core.scenarios import run
from ti_framework.io import schema as S
from ti_framework.io.csv_adapter import load_csv_inputs
from ti_framework.io.workbook import load_workbook_inputs
from ti_framework.models import EngineConfig, Powertrain, Scenario, SupportParams, Vehicle

_WORKBOOK = Path(__file__).resolve().parents[1] / "data" / "TI_Data_Workbook_v0.1.xlsx"

pytestmark = pytest.mark.skipif(not _WORKBOOK.exists(), reason="data workbook not available")


def _export_csvs(xlsx: Path, outdir: Path) -> None:
    """Dump each workbook sheet to <sheet>.csv (what an analyst's CSV export produces)."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    for name in wb.sheetnames:
        lines = []
        for row in wb[name].iter_rows(values_only=True):
            cells = []
            for c in row:
                s = "" if c is None else str(c)
                if any(ch in s for ch in ',"\n'):
                    s = '"' + s.replace('"', '""') + '"'
                cells.append(s)
            lines.append(",".join(cells))
        (outdir / f"{name}.csv").write_text("\n".join(lines), encoding="utf-8")


def test_csv_loads_same_as_xlsx(tmp_path):
    _export_csvs(_WORKBOOK, tmp_path)
    from_xlsx = load_workbook_inputs(_WORKBOOK)
    from_csv = load_csv_inputs(tmp_path)

    assert set(from_csv.countries) == set(from_xlsx.countries)
    for code, cx in from_xlsx.countries.items():
        cc = from_csv.countries[code]
        assert cc.grid_intensity == pytest.approx(cx.grid_intensity)
        assert cc.status is cx.status
        assert cc.r_fleet.s2 == pytest.approx(cx.r_fleet.s2) if cx.r_fleet.s2 is not None else cc.r_fleet.s2 is None
    assert len(from_csv.vehicles) == len(from_xlsx.vehicles)
    assert len(from_csv.volumes) == len(from_xlsx.volumes)
    assert from_csv.missing_inputs == from_xlsx.missing_inputs
    assert from_csv.warnings == from_xlsx.warnings


def test_missing_csv_files_surface_as_missing_inputs(tmp_path):
    inp = load_csv_inputs(tmp_path)  # empty directory
    assert "Registration_Vcv: no volume units collected" in inp.missing_inputs
    assert "Support_params: vehicle lifetime T not collected" in inp.missing_inputs


def _run_from_inputs(inputs, code: str) -> float:
    """Minimal end-to-end run driven by loaded country data (KR): one BEV placement."""
    country = inputs.countries[code]
    country.fleet_intensity_base = 0.165  # explicit test input — not in workbook yet
    support = SupportParams(lifetime_T=15, vkt={code: 13000.0})
    veh = Vehicle(brand="Test", model="BEV", powertrain=Powertrain.BEV, eta_ev=0.18)
    placements = [Placement(country_code=code, vehicle=veh, units=1000)]
    result = run("TestCo", 2024, placements, {code: country}, support, EngineConfig())
    return result.cohorts[Scenario.S2].total


def test_workbook_value_change_changes_output_no_code_edit(tmp_path):
    """The no-hardcode guarantee: edit one workbook cell, engine output moves."""
    wb = openpyxl.load_workbook(_WORKBOOK)
    ws = wb[S.SHEET_LAYER1]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    grid_col = header.index("Grid intensity G_c(0) gCO2/kWh (2024)") + 1
    code_col = header.index("Code") + 1
    for row in ws.iter_rows(min_row=2):
        if row[code_col - 1].value == "KR":
            original = float(row[grid_col - 1].value)
            row[grid_col - 1].value = original * 2  # double KR grid intensity
            break
    else:
        pytest.fail("KR row not found")
    modified = tmp_path / "modified.xlsx"
    wb.save(modified)

    baseline = _run_from_inputs(load_workbook_inputs(_WORKBOOK), "KR")
    changed = _run_from_inputs(load_workbook_inputs(modified), "KR")
    assert changed != pytest.approx(baseline)
    assert changed < baseline  # dirtier grid -> lower BEV contribution
