"""Extract the overseas generating units of the companies in scope from the GEM tracker workbook.

Input   data/power/projects/raw/gem_*.xlsx       HAND-GATHERED (GEM download form; see method.md)
        data/power/projects/method/gem_columns.csv           our field -> tracker header candidates
        data/power/projects/method/country_name_overrides.csv
        data/power/geography/processed/country_codes.csv
        data/power/companies/method/companies.csv            owner/parent regex and HQ per company
        data/power/roles/raw/project_roles.csv               unit/location ids named by a role
Output  data/power/projects/processed/projects_gem.csv

A unit is kept when a company pattern matches its Owner or Parent text **and the unit is outside
that company's home country** (the case study is the trade in power projects, so a Korean plant
owned by KEPCO is not in scope), or when the hand-gathered role register names its unit or
location id. Country names are mapped to alpha-2 through the tracker's own country sheet where
the workbook has one, then the geography table and the overrides; an unmapped name stops the
extractor and is listed, never dropped. The tracker's ``Type`` is normalised to the model's
``fuel_type`` (oil/gas split on the first listed fuel); the raw type is kept as ``gem_type``.

Run from the repository root:  .venv/bin/python script/power/projects/extract_gem_tracker.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "model"))
from power_io import DATA, REPO, hand_file_required, num, read_csv, write_csv  # noqa: E402
from registry import upsert_raw_file, upsert_source  # noqa: E402

DATASET = DATA / "projects"
RAW_DIR = DATASET / "raw"
COLUMNS = DATASET / "method" / "gem_columns.csv"
OVERRIDES = DATASET / "method" / "country_name_overrides.csv"
CODES = DATA / "geography" / "processed" / "country_codes.csv"
COMPANIES = DATA / "companies" / "method" / "companies.csv"
ROLES = DATA / "roles" / "raw" / "project_roles.csv"
SCOPE = DATA / "registry" / "scope.csv"
OUT = DATASET / "processed" / "projects_gem.csv"
SOURCE_ID = "gem_global_integrated_power_tracker"
BTU_TO_MJ = 1.055056e-3
HEADER_SCAN_ROWS = 6
#: The tracker's own country sheet: name column and ISO alpha-2 column headers.
COUNTRY_SHEET_NAME_HEADER = "GEM Standard Country Name/Area"
COUNTRY_SHEET_ISO_HEADER = "ISO-alpha2 Code"
#: Tracker ``Type`` -> model fuel_type; ``oil/gas`` is split on the first listed fuel.
TYPE_MAP = {
    "coal": "coal",
    "oil/gas": "",
    "gas": "gas",
    "oil": "oil",
    "nuclear": "nuclear",
    "hydropower": "hydro",
    "hydro": "hydro",
    "wind": "wind",
    "utility-scale solar": "solar",
    "solar": "solar",
    "bioenergy": "bioenergy",
    "geothermal": "geothermal",
}
LIQUID_MARKERS = ("liquid", "oil", "diesel", "crude", "kerosene", "naphtha", "mazut")
FIELDS = [
    "gem_unit_id",
    "gem_location_id",
    "country",
    "country_name",
    "plant_name",
    "unit_name",
    "gem_type",
    "fuel_type",
    "fuel_detail",
    "technology",
    "capacity_mw",
    "status",
    "start_year",
    "retired_year",
    "operator",
    "owner",
    "parent",
    "latitude",
    "longitude",
    "capacity_factor",
    "heat_rate_mj_per_kwh",
    "gem_emission_factor_kgco2_per_tj",
    "wiki_url",
    "matched_companies",
    "source_id",
    "source_file",
]
NUMERIC = {
    "capacity_mw",
    "start_year",
    "retired_year",
    "latitude",
    "longitude",
    "capacity_factor",
    "heat_rate_btu_per_kwh",
    "emission_factor_kgco2_per_tj",
}


def norm(text: object) -> str:
    """Header text lower-cased with whitespace collapsed, for tolerant matching."""
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


def map_headers(header: list[object], spec: list[dict[str, str]]) -> dict[str, int] | None:
    """Our field -> column index for one header row; None if a required field is missing."""
    lookup = {norm(h): i for i, h in enumerate(header) if h not in (None, "")}
    mapping: dict[str, int] = {}
    for s in spec:
        for candidate in s["candidates"].split(";"):
            if norm(candidate) in lookup:
                mapping[s["field"]] = lookup[norm(candidate)]
                break
        else:
            if s["required"] == "yes":
                return None
    return mapping


def company_matcher(companies: list[dict[str, str]]) -> list[tuple[str, re.Pattern[str]]]:
    """Compiled owner/parent patterns of the in-scope companies."""
    return [
        (c["company_id"], re.compile(c["gem_owner_pattern"], re.IGNORECASE))
        for c in companies
        if c["in_scope"] == "yes" and c["gem_owner_pattern"]
    ]


def match_companies(text: str, matchers: list[tuple[str, re.Pattern[str]]]) -> list[str]:
    """Company ids whose pattern occurs in the owner/parent text."""
    return [cid for cid, pattern in matchers if pattern.search(text)]


def fuel_type_of(gem_type: str, fuel_detail: str) -> str:
    """Model fuel_type from the tracker type, splitting oil/gas on the first listed fuel."""
    kind = TYPE_MAP.get(gem_type.strip().lower())
    if kind is None:
        return gem_type.strip().lower()
    if kind:
        return kind
    first = fuel_detail.split(",")[0].strip().lower()
    if "gas" in first and not any(m in first for m in LIQUID_MARKERS if m != "oil"):
        return "gas"
    if any(m in first for m in LIQUID_MARKERS):
        return "oil"
    return "gas"


def alpha2_lookup(codes: list[dict[str, str]], overrides: list[dict[str, str]]) -> dict[str, str]:
    """Lower-cased country name -> alpha-2, overrides winning over common and official names."""
    table: dict[str, str] = {}
    for r in codes:
        table[r["name_official"].lower()] = r["alpha2"]
    for r in codes:
        table[r["name_common"].lower()] = r["alpha2"]
    for r in overrides:
        table[r["tracker_name"].lower()] = r["alpha2"]
    return table


def tracker_country_sheet(wb) -> dict[str, str]:  # noqa: ANN001
    """Lower-cased tracker country name -> alpha-2 from the workbook's own country sheet."""
    out: dict[str, str] = {}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = [norm(h) for h in next(rows, [])]
        if (
            norm(COUNTRY_SHEET_NAME_HEADER) not in header
            or norm(COUNTRY_SHEET_ISO_HEADER) not in header
        ):
            continue
        i_name, i_iso = (
            header.index(norm(COUNTRY_SHEET_NAME_HEADER)),
            header.index(norm(COUNTRY_SHEET_ISO_HEADER)),
        )
        for r in rows:
            name, iso = r[i_name], r[i_iso]
            if name and iso and len(str(iso).strip()) == 2:
                out[str(name).strip().lower()] = str(iso).strip().upper()
        break
    return out


def read_sheet_rows(ws, spec: list[dict[str, str]]) -> list[dict[str, object]]:  # noqa: ANN001
    """Rows of one worksheet keyed by our field names; [] when the sheet has no unit header."""
    rows = ws.iter_rows(values_only=True)
    head: list[tuple[object, ...]] = []
    mapping = None
    for _ in range(HEADER_SCAN_ROWS):
        row = next(rows, None)
        if row is None:
            break
        head.append(row)
        mapping = map_headers(list(row), spec)
        if mapping:
            break
    if not mapping:
        return []
    out = []
    for values in rows:
        if all(v in (None, "") for v in values):
            continue
        rec: dict[str, object] = {}
        for field, i in mapping.items():
            v = values[i] if i < len(values) else None
            rec[field] = num(v) if field in NUMERIC else ("" if v is None else str(v))
        out.append(rec)
    return out


def read_scope() -> dict[str, str]:
    """The sector's scope settings (registry/scope.csv): which destinations, whether home counts."""
    return {r["setting"]: r["value"].strip() for r in read_csv(SCOPE)} if SCOPE.exists() else {}


def extract(
    workbooks: list[Path],
    spec: list[dict[str, str]],
    companies: list[dict[str, str]],
    role_ids: set[str],
    names: dict[str, str],
    scope: dict[str, str] | None = None,
) -> tuple[list[dict[str, object]], set[str], int]:
    """Kept unit rows, the country names that could not be mapped, the domestic units dropped."""
    scope = scope or {}
    exclude_home = scope.get("exclude_home_country", "yes") == "yes"
    wanted = scope.get("destinations", "all")
    destinations = None if wanted == "all" else {d.strip() for d in wanted.split(";")}
    matchers = company_matcher(companies)
    home = {c["company_id"]: c["country"] for c in companies}
    kept: list[dict[str, object]] = []
    unmapped: set[str] = set()
    seen: set[str] = set()
    domestic = 0
    for path in workbooks:
        wb = load_workbook(path, read_only=True, data_only=True)
        names_here = {**names, **tracker_country_sheet(wb)}
        for ws in wb.worksheets:
            for r in read_sheet_rows(ws, spec):
                uid = str(r.get("gem_unit_id") or "")
                lid = str(r.get("gem_location_id") or "")
                if not uid or uid in seen:
                    continue
                owner_text = f"{r.get('owner', '')} | {r.get('parent', '')}"
                matched_any = match_companies(owner_text, matchers)
                named = uid in role_ids or lid in role_ids
                if not matched_any and not named:
                    continue
                name = str(r.get("country") or "")
                alpha2 = names_here.get(name.lower())
                if alpha2 is None:
                    unmapped.add(name)
                    continue
                if destinations is not None and alpha2 not in destinations:
                    continue
                foreign = [cid for cid in matched_any if home.get(cid) != alpha2]
                if exclude_home and not foreign and not named:
                    domestic += 1
                    continue
                matched = foreign if exclude_home else matched_any
                seen.add(uid)
                heat = r.get("heat_rate_btu_per_kwh")
                gem_type = str(r.get("gem_type", "")).strip().lower()
                fuel_detail = str(r.get("fuel_detail", "") or "")
                kept.append(
                    {
                        "gem_unit_id": uid,
                        "gem_location_id": lid,
                        "country": alpha2,
                        "country_name": name,
                        "plant_name": r.get("plant_name", ""),
                        "unit_name": r.get("unit_name", ""),
                        "gem_type": gem_type,
                        "fuel_type": fuel_type_of(gem_type, fuel_detail),
                        "fuel_detail": fuel_detail,
                        "technology": r.get("technology", ""),
                        "capacity_mw": r.get("capacity_mw"),
                        "status": str(r.get("status", "")).strip().lower(),
                        "start_year": int(r["start_year"]) if r.get("start_year") else "",
                        "retired_year": int(r["retired_year"]) if r.get("retired_year") else "",
                        "operator": r.get("operator", ""),
                        "owner": r.get("owner", ""),
                        "parent": r.get("parent", ""),
                        "latitude": r.get("latitude"),
                        "longitude": r.get("longitude"),
                        "capacity_factor": r.get("capacity_factor"),
                        "heat_rate_mj_per_kwh": round(heat * BTU_TO_MJ, 4) if heat else "",
                        "gem_emission_factor_kgco2_per_tj": r.get("emission_factor_kgco2_per_tj"),
                        "wiki_url": r.get("wiki_url", ""),
                        "matched_companies": ";".join(matched),
                        "source_id": SOURCE_ID,
                        "source_file": path.name,
                    }
                )
        wb.close()
    return kept, unmapped, domestic


def role_ids_on_file() -> set[str]:
    """Unit and location ids named in the hand-gathered role register (may be empty)."""
    if not ROLES.exists():
        return set()
    ids: set[str] = set()
    for r in read_csv(ROLES):
        ids.update(x for x in (r.get("gem_unit_id", ""), r.get("gem_location_id", "")) if x)
    return ids


def main() -> None:
    """Read every tracker workbook under raw/, keep the in-scope units, register the files."""
    workbooks = sorted(p for p in RAW_DIR.glob("gem_*.xlsx") if not p.name.startswith("~$"))
    if not workbooks:
        hand_file_required(
            RAW_DIR / "gem_<tracker>_<release>.xlsx",
            "download the Global Integrated Power Tracker through GEM's form at "
            "https://globalenergymonitor.org/projects/global-integrated-power-tracker/download-data/"
            " and save it under data/power/projects/raw/ with a lowercase gem_ prefix "
            "(see data/power/projects/method/method.md)",
        )
    if not CODES.exists():
        hand_file_required(CODES, "run script/power/geography/extract_country_codes.py")
    spec = read_csv(COLUMNS)
    companies = read_csv(COMPANIES)
    names = alpha2_lookup(read_csv(CODES), read_csv(OVERRIDES))
    scope = read_scope()
    kept, unmapped, domestic = extract(workbooks, spec, companies, role_ids_on_file(), names, scope)
    if unmapped:
        raise SystemExit(
            "tracker country names with no alpha-2; add them to "
            f"{OVERRIDES.relative_to(REPO)}: {sorted(unmapped)}"
        )
    if not kept:
        raise SystemExit("no unit matched a company pattern or a role row; check gem_columns.csv")
    upsert_source(
        {
            "source_id": SOURCE_ID,
            "publisher": "Global Energy Monitor",
            "title": "Global Integrated Power Tracker (unit-level power plant data)",
            "url": "https://globalenergymonitor.org/projects/global-integrated-power-tracker/",
            "how_obtained": (
                "downloaded by hand through GEM's download form (name, organisation and email "
                "required); no scriptable link exists"
            ),
            "accessed_date": "2026-09-05",
            "license": "CC BY 4.0",
            "used_by": "projects;roles;model",
        },
        data_root=DATA,
    )
    for path in workbooks:
        upsert_raw_file(
            "projects",
            path,
            SOURCE_ID,
            path.name,
            "hand download through GEM's form; release as named in the file",
            data_root=DATA,
        )
    kept.sort(key=lambda r: (str(r["country"]), str(r["plant_name"]), str(r["gem_unit_id"])))
    write_csv(OUT, FIELDS, kept)
    by_country: dict[str, int] = {}
    for r in kept:
        by_country[str(r["country"])] = by_country.get(str(r["country"]), 0) + 1
    print(
        f"{OUT.relative_to(REPO)}: {len(kept)} overseas units in {len(by_country)} countries from "
        f"{len(workbooks)} workbook(s); {domestic} domestic units left out (scope "
        f"exclude_home_country = {scope.get('exclude_home_country', 'yes')}); by country "
        f"{dict(sorted(by_country.items(), key=lambda kv: -kv[1]))}"
    )


if __name__ == "__main__":
    main()
