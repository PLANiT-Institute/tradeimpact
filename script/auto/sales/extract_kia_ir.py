"""Extract Kia IR retail sales by model and market into the sales schema.

Reads ``data/auto/sales/raw/kia_2026_retail_sales_by_model_market.xlsx`` (Kia IR,
"Retail Sales by Country", sheet ``Total`` = year-to-date sum of the monthly sheets) and
writes ``data/auto/sales/processed/sales_kia_ir_2026.csv``.

The sheet lists models in blocks by production plant; a block ends with a subtotal row
whose label sits in column B (e.g. ``Korea Plants``). Market columns are Kia's IR regions,
so ``destination_level`` is ``region`` for Europe, Eastern Europe, Latin America, Middle
East, Africa and Asia Pacific. Labels are resolved through ``method/kia_labels.csv``.

Run from the repository root:  .venv/bin/python script/auto/sales/extract_kia_ir.py
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[3]
DATASET = REPO / "data" / "auto" / "sales"
RAW = DATASET / "raw" / "kia_2026_retail_sales_by_model_market.xlsx"
LABELS = DATASET / "method" / "kia_labels.csv"
OUT = DATASET / "processed" / "sales_kia_ir_2026.csv"

COMPANY = "kia"
BASIS = "retail_sales"
HEADER_ROW = 4
FIRST_MARKET_COL = 7  # column H; column G is the derived "Total"
MODEL_COL = 4  # column E
BLOCK_COL = 1  # column B
MONTH_SHEETS = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "June",
    "July",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]

FIELDS = [
    "company",
    "destination",
    "destination_level",
    "origin",
    "cohort_year",
    "period",
    "model",
    "powertrain",
    "units",
    "basis",
    "source_file",
]


def load_labels() -> tuple[dict[str, tuple[str, str]], dict[str, str]]:
    """Return (market label -> (code, level), plant label -> origin code)."""
    markets: dict[str, tuple[str, str]] = {}
    plants: dict[str, str] = {}
    with LABELS.open(newline="") as f:
        for row in csv.DictReader(f):
            if row["kind"] == "market":
                markets[row["label"]] = (row["code"], row["level"])
            else:
                plants[row["label"]] = row["code"]
    return markets, plants


def clean(label: object) -> str:
    """Normalise a header or row label: collapse whitespace, drop line breaks."""
    return " ".join(str(label).split())


def months_with_data(wb) -> list[int]:  # noqa: ANN001 - openpyxl workbook
    """Month numbers whose sheet carries a non-zero grand total."""
    found = []
    for i, name in enumerate(MONTH_SHEETS, start=1):
        for row in wb[name].iter_rows(values_only=True):
            if row[BLOCK_COL] and clean(row[BLOCK_COL]).lower() == "total":
                if row[6]:
                    found.append(i)
                break
    return found


def main() -> None:
    """Flatten the Total sheet to one row per model x market with units > 0."""
    markets, plants = load_labels()
    wb = load_workbook(RAW, data_only=True, read_only=True)
    ws = wb["Total"]
    rows = list(ws.iter_rows(values_only=True))

    header = rows[HEADER_ROW - 1]
    title = clean(rows[0][16])  # "Total. 2026, Kia ..." carries the year
    year = int(next(tok.strip(",") for tok in title.split() if tok.strip(",").isdigit()))
    months = months_with_data(wb)
    period = f"{year}-{months[0]:02d}..{year}-{months[-1]:02d}"

    market_cols: list[tuple[int, str, str]] = []
    for col in range(FIRST_MARKET_COL, len(header)):
        if header[col] is None:
            break
        label = clean(header[col])
        if label not in markets:
            raise SystemExit(f"unmapped market label {label!r}: add it to {LABELS.name}")
        code, level = markets[label]
        market_cols.append((col, code, level))

    out: list[dict[str, object]] = []
    pending: list[tuple[str, list[object]]] = []
    for row in rows[HEADER_ROW:]:
        block = clean(row[BLOCK_COL]) if row[BLOCK_COL] else ""
        model = clean(row[MODEL_COL]) if row[MODEL_COL] else ""
        if block.lower() == "total":
            break
        if block:
            if block not in plants:
                raise SystemExit(f"unmapped plant label {block!r}: add it to {LABELS.name}")
            origin = plants[block]
            for pending_model, values in pending:
                for col, code, level in market_cols:
                    units = values[col]
                    if units in (None, "", 0):
                        continue
                    out.append(
                        {
                            "company": COMPANY,
                            "destination": code,
                            "destination_level": level,
                            "origin": origin,
                            "cohort_year": year,
                            "period": period,
                            "model": pending_model,
                            "powertrain": "",
                            "units": int(units),
                            "basis": BASIS,
                            "source_file": RAW.name,
                        }
                    )
            pending = []
        elif model and row[6] is not None:
            pending.append((model, list(row)))

    if pending:
        raise SystemExit(f"{len(pending)} model rows without a closing plant subtotal")

    out.sort(key=lambda r: (str(r["origin"]), str(r["model"]), str(r["destination"])))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(out)

    total = sum(int(r["units"]) for r in out)
    print(f"{OUT.relative_to(REPO)}: {len(out)} rows, {total:,} units, period {period}")


if __name__ == "__main__":
    main()
