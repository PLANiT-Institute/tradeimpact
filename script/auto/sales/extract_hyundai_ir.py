"""Extract Hyundai IR overseas plant sales by model into the sales schema.

Reads ``data/auto/sales/raw/hyundai_2025_global_plant_sales.xlsx`` (Hyundai IR, "Global
Plant Sales": overseas plants only, monthly, split Domestic / Export per plant) and writes
``data/auto/sales/processed/sales_hyundai_plant_2025.csv``.

These are production-side sales (``basis = plant_sales``). Destination is known only for
the Domestic segment (the plant's own country) and for the ``Korea`` segment (exports to
Korea); Export rows carry ``destination = export`` with ``destination_level = unknown``.
Korean-built exports are not in this file. Plant codes resolve through
``method/hyundai_plant_codes.csv``.

Run from the repository root:  .venv/bin/python script/auto/sales/extract_hyundai_ir.py
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[3]
DATASET = REPO / "data" / "auto" / "sales"
RAW = DATASET / "raw" / "hyundai_2025_global_plant_sales.xlsx"
PLANTS = DATASET / "method" / "hyundai_plant_codes.csv"
OUT = DATASET / "processed" / "sales_hyundai_plant_2025.csv"

COMPANY = "hyundai"
BASIS = "plant_sales"
SHEET = "Global Plant Sales"
LABEL_COL = 1  # column B: plant code or segment
MODEL_COL = 2  # column C
FIRST_MONTH_COL = 3  # column D .. O are Jan..Dec
TOTAL_COL = 15  # column P
SEGMENTS = {"Domestic": "domestic", "Export": "export", "Korea": "korea"}
SKIP_LABELS = {"Total", "Sub-total", "Grand Total"}

FIELDS = [
    "company",
    "destination",
    "destination_level",
    "origin",
    "cohort_year",
    "period",
    "model",
    "powertrain",
    "units",
    "basis",
    "source_file",
]


def load_plants() -> dict[str, str]:
    """Return plant code -> ISO country (empty string when not a country)."""
    with PLANTS.open(newline="") as f:
        return {row["code"].strip(): row["country"] for row in csv.DictReader(f)}


def clean(label: object) -> str:
    """Normalise a label: collapse whitespace."""
    return " ".join(str(label).split())


def destination_for(segment: str, plant_country: str) -> tuple[str, str]:
    """Map a segment to (destination, destination_level)."""
    if segment == "domestic" and plant_country:
        return plant_country, "country"
    if segment == "korea":
        return "KR", "country"
    return "export", "unknown"


def main() -> None:
    """Flatten plant blocks to one row per plant x segment x model with units > 0."""
    plants = load_plants()
    ws = load_workbook(RAW, data_only=True, read_only=True)[SHEET]
    rows = list(ws.iter_rows(values_only=True))

    title = clean(rows[0][LABEL_COL])  # "Y2025 Global Plant Sales"
    year = int(next(tok for tok in title.replace("Y", " ").split() if tok.isdigit()))
    header = rows[2]
    n_months = sum(1 for c in range(FIRST_MONTH_COL, TOTAL_COL) if header[c])
    period = f"{year}-01..{year}-{n_months:02d}"

    out: list[dict[str, object]] = []
    plant = ""
    segment = ""
    mismatches = 0
    for row in rows[3:]:
        label = clean(row[LABEL_COL]) if row[LABEL_COL] else ""
        model = clean(row[MODEL_COL]) if row[MODEL_COL] else ""
        if label in SKIP_LABELS or model in SKIP_LABELS or label.startswith("*"):
            continue
        if label in SEGMENTS:
            segment = SEGMENTS[label]
        elif label:
            if label not in plants:
                raise SystemExit(f"unmapped plant code {label!r}: add it to {PLANTS.name}")
            plant, segment = label, ""
            continue
        months = [row[c] for c in range(FIRST_MONTH_COL, FIRST_MONTH_COL + n_months)]
        if not plant or all(v in (None, "") for v in months):
            continue
        units = int(sum(float(v or 0) for v in months))
        if row[TOTAL_COL] not in (None, "") and int(float(row[TOTAL_COL])) != units:
            mismatches += 1
        if units <= 0:
            continue
        destination, level = destination_for(segment, plants[plant])
        out.append(
            {
                "company": COMPANY,
                "destination": destination,
                "destination_level": level,
                "origin": plants[plant] or plant.lower(),
                "cohort_year": year,
                "period": period,
                "model": model,
                "powertrain": "",
                "units": units,
                "basis": BASIS,
                "source_file": RAW.name,
            }
        )

    out.sort(key=lambda r: (str(r["origin"]), str(r["destination"]), str(r["model"])))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(out)

    total = sum(int(r["units"]) for r in out)
    print(f"{OUT.relative_to(REPO)}: {len(out)} rows, {total:,} units, period {period}")
    if mismatches:
        print(f"warning: {mismatches} rows where the sheet's Total column != sum of months")


if __name__ == "__main__":
    main()
