"""Extract US light-duty vehicle stock and distance from FHWA Highway Statistics table VM-1.

Input   data/auto/vehicle_usage/raw/fhwa_vm1_2023.xlsx (sheet ``2023_VM-1``: annual vehicle
        distance travelled, registrations and average miles per vehicle by vehicle type, for
        the report year and the prior year)
Output  data/auto/vehicle_usage/processed/vehicle_usage_us.csv (long format)

FHWA classes light-duty vehicles by wheelbase, not by body type. ``LIGHT DUTY VEHICLES SHORT
WB`` (passenger cars, light vans, small SUVs) is the closest match to the EU passenger-car
(M1) population and is published here as ``car_stock`` / ``car_traffic``; the LONG WB class
(pickups, large SUVs, vans) is kept as separate ``ldv_long_wb_*`` series so the modeller can
widen the boundary deliberately. Miles are converted to kilometres (1 mile = 1.609344 km).

Run from the repository root:  .venv/bin/python script/auto/vehicle_usage/extract_fhwa_vm1.py
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[3]
DATASET = REPO / "data" / "auto" / "vehicle_usage"
RAW = DATASET / "raw" / "fhwa_vm1_2023.xlsx"
OUT = DATASET / "processed" / "vehicle_usage_us.csv"

SHEET = "2023_VM-1"
SOURCE_ID = "fhwa_vm1_2023"
KM_PER_MILE = 1.609344
YEAR_COL, LABEL_COL = 0, 1
# Column index (0-based) of each vehicle class in VM-1.
CLASS_COL = {"short_wb": 2, "long_wb": 5}
# Row label prefix -> (series suffix, unit after conversion, multiplier)
ITEMS = {
    "Total Rural and Urban": ("traffic", "million_vkm", KM_PER_MILE),  # millions of miles
    "Number of motor vehicles": ("stock", "vehicles", 1.0),
}
SERIES_NAME = {
    ("short_wb", "stock"): "car_stock",
    ("short_wb", "traffic"): "car_traffic",
    ("long_wb", "stock"): "ldv_long_wb_stock",
    ("long_wb", "traffic"): "ldv_long_wb_traffic",
}
FIELDS = ["country", "series", "year", "value", "unit", "source_id", "source_file"]


def main() -> None:
    """Read the report-year and prior-year rows for each item and class."""
    ws = load_workbook(RAW, data_only=True, read_only=True)[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict[str, object]] = []
    for i, row in enumerate(rows):
        label = str(row[LABEL_COL]).strip() if row[LABEL_COL] else ""
        for prefix, (suffix, unit, mult) in ITEMS.items():
            if not label.startswith(prefix):
                continue
            # The report-year row carries the label; the prior-year row follows it unlabelled.
            for r in (row, rows[i + 1]):
                year = int(str(r[YEAR_COL]).strip())
                for cls, col in CLASS_COL.items():
                    value = r[col]
                    if value is None:
                        continue
                    out.append(
                        {
                            "country": "US",
                            "series": SERIES_NAME[(cls, suffix)],
                            "year": year,
                            "value": round(float(value) * mult, 3),
                            "unit": unit,
                            "source_id": SOURCE_ID,
                            "source_file": RAW.name,
                        }
                    )
    if len(out) != len(ITEMS) * len(CLASS_COL) * 2:
        raise SystemExit(f"expected {len(ITEMS) * len(CLASS_COL) * 2} rows, parsed {len(out)}")
    out.sort(key=lambda r: (str(r["series"]), int(str(r["year"]))))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(out)
    latest = max(int(str(r["year"])) for r in out)
    stock = float(
        str(next(r["value"] for r in out if r["series"] == "car_stock" and r["year"] == latest))
    )
    traffic = float(
        str(next(r["value"] for r in out if r["series"] == "car_traffic" and r["year"] == latest))
    )
    print(
        f"{OUT.relative_to(REPO)}: {len(out)} rows; US short-WB {latest}: "
        f"{stock:,.0f} vehicles, {traffic * 1e6 / stock:,.0f} km/vehicle/yr"
    )


if __name__ == "__main__":
    main()
