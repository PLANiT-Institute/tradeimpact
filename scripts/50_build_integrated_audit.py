#!/usr/bin/env python3
"""Build TI_통합파일.xlsx — the team-facing integrated audit workbook.

Full-value dump of every number behind the published TI dataset, plus a live
verification gate. Never edit the xlsx by hand; regenerate with:

    ti-framework/.venv/bin/python scripts/50_build_integrated_audit.py

Fails closed if any input file is missing. Exit code 1 if any check FAILs.
"""

from __future__ import annotations

import json
import subprocess
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "ti-framework"))
sys.path.insert(0, str(REPO / "data-pipeline"))

PUBLISHED = REPO / "data" / "published"
WORKBOOK = REPO / "ti-framework" / "data" / "TI_Data_Workbook_v0.1.xlsx"
OUT = REPO / "TI_통합파일.xlsx"

REQUIRED = [
    PUBLISHED / "meta.json",
    PUBLISHED / "firms.json",
    PUBLISHED / "countries.json",
    PUBLISHED / "contract.json",
    WORKBOOK,
    REPO / "data-pipeline" / "ESTIMATES.md",
    REPO / "data-pipeline" / "COLLECTION_STATUS.md",
    REPO / "data-pipeline" / "SECTORAL_SOURCES.md",
]

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FFC7CE")
HEADER = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=13)
BASE_FONT = Font(name="Arial")


def fail_closed() -> None:
    missing = [p for p in REQUIRED if not p.exists()]
    if missing:
        print("빌드 중단 — 입력 파일 누락 (재발행 필요: data-pipeline/build_dataset.py):")
        for p in missing:
            print(f"  - {p.relative_to(REPO)}")
        raise SystemExit(1)


def style_sheet(ws, header_row: int = 1, widths: dict[int, int] | None = None) -> None:
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.fill = HEADER
            cell.font = HEADER_FONT
    for col in range(1, ws.max_column + 1):
        width = (widths or {}).get(col)
        if width is None:
            longest = max(
                (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200) + 1)),
                default=8,
            )
            width = min(max(longest + 2, 10), 55)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows():
        longest = 0
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            longest = max(longest, len(str(cell.value or "")))
        ws.row_dimensions[row[0].row].height = 75 if longest > 120 else (45 if longest > 60 else 22)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_rows(ws, header: list[str], rows: list[list], start: int = 1) -> None:
    ws.append(header) if start == 1 else None
    if start != 1:
        for j, h in enumerate(header, 1):
            ws.cell(row=start, column=j, value=h)
    for row in rows:
        ws.append([("" if v is None else v) for v in row])
    for row in ws.iter_rows():
        for cell in row:
            if cell.font == Font():
                cell.font = BASE_FONT


def run_cmd(cmd: list[str], cwd: Path) -> tuple[str, str]:
    """Return (PASS/FAIL/SKIP, last meaningful output line)."""
    try:
        proc = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True, timeout=600)
    except FileNotFoundError:
        return "SKIP", f"명령 없음: {cmd[0]}"
    except subprocess.TimeoutExpired:
        return "FAIL", "timeout"
    out = (proc.stdout + proc.stderr).strip().splitlines()
    tail = out[-1] if out else ""
    return ("PASS" if proc.returncode == 0 else "FAIL"), tail


def check_publication_gate() -> tuple[str, str]:
    """Fail if an estimated/illustrative firm report leaks into the public inventory."""
    firms = json.loads((PUBLISHED / "firms.json").read_text())
    countries = json.loads((PUBLISHED / "countries.json").read_text())
    contract = json.loads((PUBLISHED / "contract.json").read_text())
    expected = {"contract.json", "countries.json", "firms.json", "meta.json"}
    actual = {path.name for path in PUBLISHED.glob("*.json")}
    problems = []
    if actual != expected:
        problems.append(f"발행 파일={sorted(actual)}")
    if any(firm.get("runnable") for firm in firms):
        problems.append("runnable 기업 존재")
    if any("fleet_intensity_base" in country or "vkt" in country for country in countries):
        problems.append("국가 계약에 차량 추정 필드 존재")
    if "by_year" in contract.get("firm_result", []):
        problems.append("과거 시계열 계약 존재")
    if problems:
        return "FAIL", "; ".join(problems)
    return "PASS", "기업 추정 보고서 0건 · 차량 추정 필드 없음 · 과거 시계열 없음"


def parse_md_tables(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    """Return (section, header, rows) for every markdown table in the file."""
    tables = []
    section = path.name
    lines = path.read_text().splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("#"):
            section = line.lstrip("#").strip()
        if line.lstrip().startswith("|") and i + 1 < len(lines) and set(lines[i + 1].replace("|", "").replace(":", "").strip()) <= {"-", " "}:
            header = [c.strip() for c in line.strip().strip("|").split("|")]
            rows = []
            i += 2
            while i < len(lines) and lines[i].lstrip().startswith("|"):
                rows.append([c.strip() for c in lines[i].strip().strip("|").split("|")])
                i += 1
            tables.append((section, header, rows))
            continue
        i += 1
    return tables


def flat_country(code: str, c: dict) -> list:
    return [
        code,
        c.get("name"),
        c.get("status"),
        c.get("tier"),
        c.get("flag_reason"),
        c.get("grid_intensity"),
        c["r_fleet"].get("s1"),
        c["r_fleet"].get("s2"),
        c["r_fleet"].get("s2_upper"),
        c["r_fleet"].get("s3"),
        c["r_power"].get("s1"),
        c["r_power"].get("s2"),
        c["r_power"].get("s2_upper"),
        c["r_power"].get("s3"),
        "; ".join(c.get("warnings") or []),
        c.get("source"),
    ]


COUNTRY_HEADER = [
    "코드", "국가", "상태", "Tier", "FLAG 사유", "그리드 원단위 (kgCO2e/kWh)",
    "r_fleet S1", "r_fleet S2", "r_fleet S2상단", "r_fleet S3",
    "r_power S1", "r_power S2", "r_power S2상단", "r_power S3",
    "경고", "출처",
]

def main() -> int:
    fail_closed()

    meta = json.loads((PUBLISHED / "meta.json").read_text())
    firms = json.loads((PUBLISHED / "firms.json").read_text())
    countries = json.loads((PUBLISHED / "countries.json").read_text())

    py = REPO / "ti-framework" / ".venv" / "bin" / "python"
    checks = [
        ("발행 데이터 전량 재계산 (check_published.py)",
         *run_cmd([str(py), "data-pipeline/check_published.py"], REPO)),
        ("이론↔코드↔테스트 동기화 (check_sync.py)",
         *run_cmd([str(py), "scripts/check_sync.py"], REPO)),
        ("엔진 테스트 스위트 (pytest)",
         *run_cmd([str(py), "-m", "pytest", "-q"], REPO / "ti-framework")),
        ("추정·시계열 공개 차단 (이 스크립트 내장 게이트)",
         *check_publication_gate()),
        ("웹 데이터 계약 테스트 (npm test)",
         *run_cmd(["npm", "test"], REPO / "web")),
    ]

    wb = Workbook()

    # 00_개요
    ws = wb.active
    ws.title = "00_개요"
    overview = [
        ["TI 통합 파일 (Trade Impact — 통합 데이터 감사 워크북)", ""],
        ["", ""],
        ["생성일", date.today().isoformat()],
        ["생성 스크립트", "scripts/50_build_integrated_audit.py (수기 편집 금지 — 스크립트로만 재생성)"],
        ["재생성 명령", "ti-framework/.venv/bin/python scripts/50_build_integrated_audit.py"],
        ["대상 데이터셋", f"data/published/ (engine {meta['engine_version']}, workbook {meta['workbook']})"],
        ["", ""],
        ["이 파일은 무엇인가", "발행된 소스 기반 국가 벤치마크, 원본 워크북, 출처, 수집 백로그, 검증 결과를 한 파일에 수록한 감사용 스냅샷. 회사별 추정 결과와 과거 시계열은 포함하지 않음."],
        ["색상 범례", "초록 = 기계 검증 완료 · 노랑 = 작성자/팀 확인 필요 · 빨강 = 검증 실패"],
        ["", ""],
        ["시트 안내", ""],
        ["01_검증결과", "빌드 시점에 실제 실행한 검증 게이트 결과"],
        ["02_해시증빙", "발행본 무결성 해시 (엔진 소스, 워크북, 입력, 데이터셋)"],
        ["03_국가벤치마크", "11개국 NDC 벤치마크 전량 (출처·Tier·경고 포함)"],
        ["04_지원파라미터", "부문 공통 파라미터 (VKT, 수명, UF 밴드, 실주행 보정)"],
        ["05_기업유니버스", "TI/CAP 대상기업 전체 목록과 실행 가능 여부"],
        ["06_제거기록", "ESTIMATES.md — 제거된 추정 입력과 재게시 증거 게이트"],
        ["07_수집백로그", "COLLECTION_STATUS.md — 미수집 항목과 데이터 경고"],
        ["08_부문별출처", "SECTORAL_SOURCES.md — 국가 경로·차량 인증 파라미터 출처"],
        ["WB_*", "원본 데이터 워크북 시트 전량 덤프"],
        ["99_체크리스트", "팀 확인란 (노랑 칸에 확인일·서명 기입)"],
        ["", ""],
        ["내부 검증 fixture", "ReferenceCo는 엔진 산술 단위검증에만 사용하며 이 감사 파일과 공개 데이터에 포함하지 않음"],
    ]
    for row in overview:
        ws.append(row)
    ws["A1"].font = Font(name="Arial", bold=True, size=15)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110

    # 01_검증결과
    ws = wb.create_sheet("01_검증결과")
    write_rows(ws, ["검증 게이트", "결과", "상세 (마지막 출력 라인)"],
               [[name, status, detail] for name, status, detail in checks])
    for i, (_, status, _) in enumerate(checks, 2):
        fill = GREEN if status == "PASS" else (YELLOW if status == "SKIP" else RED)
        for j in range(1, 4):
            ws.cell(row=i, column=j).fill = fill
    style_sheet(ws, widths={1: 45, 2: 8, 3: 90})

    # 02_해시증빙
    ws = wb.create_sheet("02_해시증빙")
    prov_rows = [
        ["engine_version", meta["engine_version"], "엔진 패키지 버전"],
        ["engine_source_sha256", meta["engine_source_sha256"], "ti_framework 소스 트리 해시"],
        ["workbook", meta["workbook"], "원본 데이터 워크북 파일명"],
        ["workbook_sha256", meta["workbook_sha256"], "워크북 파일 해시"],
        ["dataset_sha256", meta["dataset_sha256"], "발행 데이터셋 전체 해시"],
        ["generated_utc", meta.get("generated_utc"), "발행 시각 (UTC)"],
    ] + [[f"target_sources_sha256.{k}", v, "대상기업 원본 워크북 해시"]
         for k, v in meta.get("target_sources_sha256", {}).items()]
    write_rows(ws, ["항목", "값", "설명"], prov_rows)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = GREEN
    style_sheet(ws, widths={1: 45, 2: 70, 3: 40})

    # 03_국가벤치마크
    ws = wb.create_sheet("03_국가벤치마크")
    write_rows(ws, COUNTRY_HEADER, [flat_country(c["code"], c) for c in countries])
    style_sheet(ws)

    # 04_지원파라미터
    ws = wb.create_sheet("04_지원파라미터")
    sc = meta["support_contract"]
    rows = [
        ["차량 수명 T (년)", sc["lifetime_T"], "전 시장 공통, ±민감도"],
        ["수명 민감도 ± (년)", sc["lifetime_sens"], ""],
        ["실주행 보정 범위", sc["realworld_range"], "인증치 대비 배수"],
        ["PHEV UF 하한 밴드", sc["uf_band"], "규제 UF − 밴드 = 하한 병기 (Guideline §3.5)"],
    ] + [[f"VKT — {code}", v, "km/년"] for code, v in sorted(sc["vkt"].items())]
    write_rows(ws, ["파라미터", "값", "비고"], rows)
    style_sheet(ws, widths={1: 30, 2: 20, 3: 50})

    # 05_기업유니버스
    ws = wb.create_sheet("05_기업유니버스")
    write_rows(
        ws,
        ["기업명", "슬러그", "프로젝트", "섹터", "국가", "실행 가능", "예시용", "코호트 연도", "비고"],
        [[f.get("name"), f.get("slug"), f.get("project"), f.get("sector"), f.get("country"),
          f.get("runnable"), f.get("illustrative", False), f.get("cohort_year"), f.get("note")]
         for f in firms],
    )
    style_sheet(ws)

    # 제거 기록 / 수집 백로그 / 출처
    for sheet_name, md in (
        ("06_제거기록", REPO / "data-pipeline" / "ESTIMATES.md"),
        ("07_수집백로그", REPO / "data-pipeline" / "COLLECTION_STATUS.md"),
        ("08_부문별출처", REPO / "data-pipeline" / "SECTORAL_SOURCES.md"),
    ):
        ws = wb.create_sheet(sheet_name)
        r = 1
        ws.cell(row=r, column=1, value=f"원문: {md.relative_to(REPO)} (전체 표 자동 추출)").font = TITLE_FONT
        r += 2
        for section, header, rows in parse_md_tables(md):
            ws.cell(row=r, column=1, value=section).font = Font(name="Arial", bold=True, size=11)
            r += 1
            for j, h in enumerate(header, 1):
                c = ws.cell(row=r, column=j, value=h)
                c.fill = HEADER
                c.font = HEADER_FONT
            r += 1
            for row in rows:
                for j, v in enumerate(row, 1):
                    ws.cell(row=r, column=j, value=v).font = BASE_FONT
                r += 1
            r += 1
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 40
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row[0].row].height = 45
        if sheet_name == "07_수집백로그":
            ws.cell(row=r, column=1, value="엔진 보고 누락 입력 (meta.json collection_status)").font = Font(name="Arial", bold=True, size=11)
            r += 1
            for item in meta["collection_status"]["missing_inputs"]:
                cell = ws.cell(row=r, column=1, value=item)
                cell.font = BASE_FONT
                cell.fill = YELLOW
                r += 1
            r += 1
            ws.cell(row=r, column=1, value="엔진 보고 경고").font = Font(name="Arial", bold=True, size=11)
            r += 1
            for item in meta["collection_status"]["warnings"]:
                cell = ws.cell(row=r, column=1, value=item)
                cell.font = BASE_FONT
                cell.fill = YELLOW
                r += 1

    # WB_* 원본 워크북 덤프
    src = load_workbook(WORKBOOK, data_only=True)
    for sheet in src.worksheets:
        ws = wb.create_sheet(f"WB_{sheet.title}"[:31])
        for row in sheet.iter_rows(values_only=True):
            ws.append(list(row))
        if ws.max_row >= 1:
            style_sheet(ws)

    # 99_체크리스트
    ws = wb.create_sheet("99_체크리스트")
    checklist = [
        "01_검증결과 시트의 모든 게이트가 PASS인지 확인",
        "03_국가벤치마크 출처(UNFCCC NDC, Ember 2024)가 최신인지 확인",
        "06_제거기록의 추정 fixture·과거 시계열 삭제 범위 확인",
        "07_수집백로그 항목별 담당자 지정",
        "차량 인증행을 회사 평균으로 해석하지 않는지 확인",
        "모든 기업이 runnable=false이고 공개 보고서 JSON이 없는지 확인",
        "웹 국가 페이지가 회사 점수 대신 출처·경고만 표시하는지 확인",
    ]
    write_rows(ws, ["항목", "확인자", "확인일", "비고"], [[c, "", "", ""] for c in checklist])
    for i in range(2, len(checklist) + 2):
        for j in (2, 3, 4):
            ws.cell(row=i, column=j).fill = YELLOW
    style_sheet(ws, widths={1: 70, 2: 15, 3: 15, 4: 40})

    wb.save(OUT)

    failed = [name for name, status, _ in checks if status == "FAIL"]
    print(f"저장: {OUT.relative_to(REPO)} ({len(wb.sheetnames)} 시트)")
    for name, status, detail in checks:
        print(f"  [{status}] {name} — {detail}")
    if failed:
        print("검증 실패 게이트 존재 — 워크북의 01_검증결과 시트 확인")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
