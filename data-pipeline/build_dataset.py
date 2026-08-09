#!/usr/bin/env python3
# SPDX-License-Identifier: GPL-3.0-or-later
"""Build the published dataset the web app and MCP server query.

    workbook.xlsx + firm universe + sector registry
        -> data/published/{firms,countries,sectors,benchmarks,company_metrics,sources,...}.json

Observed product cohorts, destination pathways, and calculation readiness are published as
separate objects. A lifetime result is not published until activity, product, destination-use,
survival, energy, and scenario inputs pass the evidence gate. The illustrative reference fixture
is used only to pin the engine schema; it is never emitted as a public assessment.

Usage:  ti-framework/.venv/bin/python data-pipeline/build_dataset.py
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from importlib.metadata import version as pkg_version
from math import isfinite
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
ENGINE_DIR = REPO / "ti-framework"
WORKBOOK = ENGINE_DIR / "data" / "TI_Data_Workbook_v0.1.xlsx"
OUT = REPO / "data" / "published"

sys.path.insert(0, str(ENGINE_DIR))

from ti_framework.core.scenarios import run  # noqa: E402
from ti_framework.core.sensitivity import run_sensitivity  # noqa: E402
from ti_framework.alignment.registry import list_sector_profiles  # noqa: E402
from ti_framework.io.fixtures import load_fixture  # noqa: E402
from ti_framework.io.schema import FLAG_REASONS  # noqa: E402
from ti_framework.io.workbook import load_workbook_inputs  # noqa: E402
from ti_framework.models import BenchmarkStatus  # noqa: E402
from ti_framework.report.outputs import to_json_dict  # noqa: E402

from adapters.automotive_eea import HYUNDAI_SNAPSHOT, TOYOTA_SNAPSHOT  # noqa: E402
from adapters.automotive_eea import build_all_records as build_eea_automotive_records  # noqa: E402
from adapters.destination_eu import SNAPSHOT as DESTINATION_SNAPSHOT  # noqa: E402
from adapters.destination_eu import build_records as build_destination_records  # noqa: E402
from lifetime_run import build_fixtures, run_lifetime  # noqa: E402
from adapters.power_jera import SNAPSHOT as JERA_POWER_SNAPSHOT  # noqa: E402
from adapters.power_jera import build_records as build_jera_power_records  # noqa: E402
from adapters.power_koen import SNAPSHOT as KOEN_POWER_SNAPSHOT  # noqa: E402
from adapters.power_koen import build_records as build_koen_power_records  # noqa: E402
from adapters.shipping_mol import SNAPSHOT as MOL_SHIPPING_SNAPSHOT  # noqa: E402
from adapters.shipping_mol import build_records as build_mol_shipping_records  # noqa: E402

VALIDATION_FIXTURE = ENGINE_DIR / "fixtures" / "reference_case.json"

ASSESSMENT_GATE_NOTE = (
    "Company assessment withheld: source-backed country/year/model/powertrain registration "
    "volumes and the remaining calculation inputs have not been collected. Estimated vehicle "
    "mix and reconstructed historical cohorts were removed on 2026-08-03."
)
TOYOTA_ALIGNMENT_NOTE = (
    "Observed Toyota-brand 2024 EU27 destination cohort available by country, commercial name, "
    "and powertrain. Lifetime TI remains withheld until destination-use and pathway inputs are "
    "source-complete."
)
HYUNDAI_ALIGNMENT_NOTE = (
    "Observed Hyundai-brand 2024 EU27 destination cohort available by country, commercial name, "
    "and powertrain. Lifetime TI and production-to-destination origin mapping remain withheld."
)
JERA_ALIGNMENT_NOTE = (
    "Evidence-first FY2024 Japan power snapshot available from independently assured company "
    "data. National 2030/2040 targets are context only because their boundaries do not match."
)
KOEN_ALIGNMENT_NOTE = (
    "Evidence-first 2024 Korea power snapshot available from KOEN's official ESG table. "
    "Generation basis, plant-total reconciliation, and assurance limitations prevent an "
    "emissions-intensity or target-gap calculation."
)
MOL_ALIGNMENT_NOTE = (
    "Evidence-first FY2024 global-shipping snapshot available from independently assured MOL "
    "EEOI data. IMO 2030 ambitions are context only because baseline, GHG scope, and aggregation "
    "method do not match."
)

RATE_FIELDS = ("s1", "s2", "s3", "s2_upper")

def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _file_sha256(path: Path) -> str:
    return _sha256_bytes(path.read_bytes())


# Eight digits is five orders of magnitude finer than any Tier A input here, and far coarser
# than the ~1e-14 relative spread that accumulating 1,286 terms produces across platforms.
PUBLISHED_SIGNIFICANT_DIGITS = 8


def canonical_floats(value: object) -> object:
    """Round every float to a fixed significance before it is hashed or written.

    ``pow`` and ``exp`` are allowed to differ in their last bits between platforms, so a full
    17-digit float makes the published artifact machine-dependent for no analytical gain — none
    of these inputs carry twelve significant digits, let alone seventeen.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if not isfinite(value):
            raise ValueError(f"refusing to publish a non-finite value: {value!r}")
        return float(f"{value:.{PUBLISHED_SIGNIFICANT_DIGITS}g}")
    if isinstance(value, dict):
        return {key: canonical_floats(item) for key, item in value.items()}
    if isinstance(value, list):
        return [canonical_floats(item) for item in value]
    return value


def _json_sha256(value: object) -> str:
    canonical = json.dumps(
        canonical_floats(value), sort_keys=True, separators=(",", ":"), ensure_ascii=False
    )
    return _sha256_bytes(canonical.encode())


def _tree_sha256(paths: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths):
        digest.update(str(path.relative_to(REPO)).encode())
        digest.update(b"\0")
        digest.update(path.read_bytes())
        digest.update(b"\0")
    return digest.hexdigest()


def _write_json(path: Path, value: object) -> None:
    """One canonical JSON writer for reviewable, reproducible published artifacts."""
    path.write_text(
        json.dumps(canonical_floats(value), indent=2, ensure_ascii=False, sort_keys=True) + "\n"
    )


def _prune_stale_json_outputs(expected_names: set[str]) -> list[str]:
    """Remove report JSON that is no longer part of the successfully built inventory."""
    removed: list[str] = []
    for path in OUT.glob("*.json"):
        if path.name not in expected_names:
            path.unlink()
            removed.append(path.name)
    return sorted(removed)


def slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def _rows(path: Path, expect_first_header: str) -> list[dict]:
    """Read a target-companies sheet: find the header row, return data-row dicts."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(i for i, r in enumerate(rows) if r[0] == expect_first_header)
    header = [str(c) if c else "" for c in rows[header_i]]
    out = []
    for r in rows[header_i + 1 :]:
        if not r[0] or not r[2]:  # needs Sector and Company columns
            continue
        out.append(dict(zip(header, (str(c) if c is not None else "" for c in r), strict=False)))
    return out


def build_universe() -> list[dict]:
    firms: list[dict] = []
    for r in _rows(REPO / "TI_CaseStudy_Target_Companies.xlsx", "Sector"):
        name = r["Company (candidate)"]
        slug = slugify(name)
        if slug == "mitsui" and r["Sector"] == "Shipping":
            name = "Mitsui O.S.K. Lines (MOL)"
        firms.append(
            {
                "slug": slug,
                "name": name,
                "sector": r["Sector"],
                "country": r["Country"],
                "project": "TI",
                "runnable": False,
                "note": ASSESSMENT_GATE_NOTE if slug in {"toyota", "hyundai"} else None,
                "status": r.get("Status", ""),
                "selection_criteria": r.get("Selection criteria", ""),
            }
        )
    for r in _rows(REPO / "CAP_Target_Companies_Draft.xlsx", "Sector"):
        name = r["Company (candidate)"]
        firms.append(
            {
                "slug": slugify(name),
                "name": name,
                "sector": r["Sector"],
                "country": r["Country"],
                "project": "CAP",
                "runnable": False,
                "ticker": r.get("Ticker/Listing", ""),
            }
        )
    if not any(firm["slug"] == "koen" for firm in firms):
        firms.append(
            {
                "slug": "koen",
                "name": "Korea South-East Power (KOEN)",
                "sector": "Power",
                "country": "Korea",
                "project": "TI",
                "runnable": False,
                "note": None,
                "status": "Evidence pilot",
                "selection_criteria": (
                    "Second power company and second operating geography with official "
                    "generation and plant-level emissions disclosure"
                ),
            }
        )
    return firms


def build_meta() -> dict:
    wb_inputs = load_workbook_inputs(WORKBOOK)
    engine_paths = list((ENGINE_DIR / "ti_framework").rglob("*.py")) + [
        ENGINE_DIR / "pyproject.toml"
    ]
    return {
        "engine": "ti_framework",
        "engine_version": pkg_version("ti-framework"),
        "engine_source_sha256": _tree_sha256(engine_paths),
        "workbook": WORKBOOK.name,
        "workbook_sha256": _file_sha256(WORKBOOK),
        "target_sources_sha256": {
            "TI_CaseStudy_Target_Companies.xlsx": _file_sha256(
                REPO / "TI_CaseStudy_Target_Companies.xlsx"
            ),
            "CAP_Target_Companies_Draft.xlsx": _file_sha256(
                REPO / "CAP_Target_Companies_Draft.xlsx"
            ),
        },
        "alignment_inputs_sha256": {
            str(TOYOTA_SNAPSHOT.relative_to(REPO)): _file_sha256(TOYOTA_SNAPSHOT),
            str(HYUNDAI_SNAPSHOT.relative_to(REPO)): _file_sha256(HYUNDAI_SNAPSHOT),
            "data-pipeline/adapters/automotive_eea.py": _file_sha256(
                REPO / "data-pipeline" / "adapters" / "automotive_eea.py"
            ),
            str(JERA_POWER_SNAPSHOT.relative_to(REPO)): _file_sha256(JERA_POWER_SNAPSHOT),
            "data-pipeline/adapters/power_jera.py": _file_sha256(
                REPO / "data-pipeline" / "adapters" / "power_jera.py"
            ),
            str(KOEN_POWER_SNAPSHOT.relative_to(REPO)): _file_sha256(KOEN_POWER_SNAPSHOT),
            "data-pipeline/adapters/power_koen.py": _file_sha256(
                REPO / "data-pipeline" / "adapters" / "power_koen.py"
            ),
            str(MOL_SHIPPING_SNAPSHOT.relative_to(REPO)): _file_sha256(MOL_SHIPPING_SNAPSHOT),
            "data-pipeline/adapters/shipping_mol.py": _file_sha256(
                REPO / "data-pipeline" / "adapters" / "shipping_mol.py"
            ),
            str(DESTINATION_SNAPSHOT.relative_to(REPO)): _file_sha256(DESTINATION_SNAPSHOT),
            "data-pipeline/adapters/destination_eu.py": _file_sha256(
                REPO / "data-pipeline" / "adapters" / "destination_eu.py"
            ),
            "data-pipeline/lifetime_run.py": _file_sha256(REPO / "data-pipeline" / "lifetime_run.py"),
        },
        "collection_status": {
            "countries_loaded": len(wb_inputs.countries),
            "missing_inputs": wb_inputs.missing_inputs,
            "warnings": wb_inputs.warnings,
        },
    }


def effective_inputs(fx, fixture_path: Path) -> dict:
    """Serialise internal validation inputs used only to pin the report schema."""
    raw = json.loads(fixture_path.read_text())
    raw.pop("expected", None)
    for code, c in fx.countries.items():
        target = raw["countries"][code]
        if c.grid_intensity is not None:
            target["grid_intensity"] = c.grid_intensity
        target["status"] = c.status.value
        target["tier"] = c.tier.value
        for attr in ("r_fleet", "r_power"):
            rate = getattr(c, attr)
            target[attr] = {
                name: value for name in RATE_FIELDS if (value := getattr(rate, name)) is not None
            }
    for target, placement in zip(raw.get("placements", []), fx.placements, strict=True):
        target.setdefault("vehicle_tier", placement.vehicle.tier.value)
        target.setdefault("volume_tier", placement.volume_tier.value)
        if placement.vehicle.source:
            target.setdefault("vehicle_source", placement.vehicle.source)
        if placement.volume_source:
            target.setdefault("volume_source", placement.volume_source)
    return raw


def _published_rates(rate) -> dict[str, float]:
    return {
        key: value
        for key in RATE_FIELDS
        if (value := getattr(rate, key)) is not None
    }


def build_countries() -> tuple[list[dict], dict]:
    """Publish only workbook-backed operating-country benchmark fields.

    Vehicle fleet baselines, VKT, lifetime, utility-factor bands, and other support
    defaults are deliberately omitted/null until their workbook cells are sourced.
    """
    inputs = load_workbook_inputs(WORKBOOK)
    countries = []
    for code, country in sorted(inputs.countries.items()):
        status = BenchmarkStatus(country.status)
        countries.append(
            {
                "code": code,
                "name": country.name,
                "grid_intensity": country.grid_intensity,
                "r_fleet": _published_rates(country.r_fleet),
                "r_power": _published_rates(country.r_power),
                "status": status.value,
                "tier": country.tier.value,
                "source": country.source,
                "warnings": country.warnings,
                "flag_reason": FLAG_REASONS.get(status) if status.is_flag else None,
            }
        )
    support_contract = {
        "lifetime_T": inputs.support.lifetime_T,
        "lifetime_sens": None,
        "uf_band": None,
        "realworld_range": None,
        "vkt": {},
    }
    return countries, support_contract


def build_alignment_data() -> dict[str, list[dict]]:
    """Build the shared evidence and exported-product contracts without inventing inputs.

    Sector adapters may add records only after source, scope, unit, and mapping coverage pass
    their validation. No adapter may fill unmatched activity with an estimated mix.
    """
    destination = build_destination_records()
    destination_rows = destination["destination_inputs"]
    # Readiness is decided by what the join actually resolved, so the fixtures are built first.
    coverage = {
        cohort_id: fixture["coverage"]
        for cohort_id, fixture in build_fixtures(destination=destination_rows).items()
    }
    automotive = build_eea_automotive_records(destination_rows, coverage)
    adapters = [
        automotive,
        {"benchmarks": [], "company_metrics": [], "sources": destination["sources"]},
        build_jera_power_records(),
        build_koen_power_records(),
        build_mol_shipping_records(),
    ]
    result = {
        "sectors": list_sector_profiles(),
        "benchmarks": [row for adapter in adapters for row in adapter["benchmarks"]],
        "company_metrics": [row for adapter in adapters for row in adapter["company_metrics"]],
        "sources": [row for adapter in adapters for row in adapter["sources"]],
        "product_cohorts": automotive["product_cohorts"],
        "pathways": automotive["pathways"],
        "impact_readiness": automotive["impact_readiness"],
        "destination_inputs": destination_rows,
    }
    _require_unique(result["sources"], "source_id")
    _require_unique(result["benchmarks"], "benchmark_id")
    metric_keys = [
        (
            row["company_id"],
            row["sector"],
            row["geography"],
            row["observation_year"],
            row["metric_id"],
            json.dumps(row.get("scope", {}), sort_keys=True),
        )
        for row in result["company_metrics"]
    ]
    if len(metric_keys) != len(set(metric_keys)):
        raise ValueError("duplicate alignment company metric identity")
    _validate_alignment_data(result)
    _validate_impact_data(result)
    return result


def _validate_impact_data(data: dict[str, list[dict]]) -> None:
    """Fail publication when a cohort, pathway, or readiness record loses provenance."""
    source_ids = {row["source_id"] for row in data["sources"]}
    cohort_ids: set[str] = set()
    for cohort in data["product_cohorts"]:
        cohort_id = cohort["cohort_id"]
        if cohort_id in cohort_ids:
            raise ValueError(f"duplicate product cohort {cohort_id}")
        cohort_ids.add(cohort_id)
        if set(cohort.get("source_ids", [])) - source_ids:
            raise ValueError(f"product cohort {cohort_id} has invalid sources")
        origin_context = cohort.get("origin_context", {})
        if set(origin_context.get("source_ids", [])) - source_ids:
            raise ValueError(f"product cohort {cohort_id} has invalid origin-context sources")
        records = cohort.get("records", [])
        if not records or sum(row["units"] for row in records) != cohort["coverage"]["reported_units"]:
            raise ValueError(f"product cohort {cohort_id} does not reconcile")
        if any(not row.get("destination_geography") or not row.get("product_type") for row in records):
            raise ValueError(f"product cohort {cohort_id} has an unmapped destination/product type")
        if any(set(row.get("source_ids", [])) - source_ids for row in records):
            raise ValueError(f"product cohort {cohort_id} record has invalid sources")
    for pathway in data["pathways"]:
        if set(pathway.get("source_ids", [])) - source_ids or not pathway.get("source_ids"):
            raise ValueError(f"pathway {pathway['pathway_id']} has invalid sources")
        if pathway.get("comparison_role") not in {"sector_proxy", "fallback_context"}:
            raise ValueError(f"pathway {pathway['pathway_id']} has invalid comparison role")
    for readiness in data["impact_readiness"]:
        if readiness.get("cohort_id") not in cohort_ids:
            raise ValueError("impact readiness references an unknown cohort")
        if readiness.get("status") == "available" and readiness.get("missing_required_inputs"):
            raise ValueError("an available impact result cannot retain required missing inputs")


def _require_unique(rows: list[dict], key: str) -> None:
    values = [row[key] for row in rows]
    if len(values) != len(set(values)):
        raise ValueError(f"duplicate alignment {key}")


def _validate_alignment_data(alignment: dict[str, list[dict]]) -> None:
    """Fail publication when evidence links or comparison contracts are inconsistent."""
    source_ids = {row["source_id"] for row in alignment["sources"]}
    sector_metrics: dict[str, dict[str, str]] = {}
    sector_direct_metrics: dict[str, dict[str, str]] = {}
    for sector in alignment["sectors"]:
        direct = {row["metric_id"]: row["unit"] for row in sector["direct_metrics"]}
        descriptive = {
            row["metric_id"]: row["unit"] for row in sector["descriptive_metrics"]
        }
        if set(direct) & set(descriptive):
            raise ValueError(f"sector {sector['sector_id']} repeats a metric definition")
        sector_direct_metrics[sector["sector_id"]] = direct
        sector_metrics[sector["sector_id"]] = {**direct, **descriptive}

    for source in alignment["sources"]:
        if not str(source.get("url", "")).startswith("https://"):
            raise ValueError(f"source {source['source_id']} must use an HTTPS URL")

    for metric in alignment["company_metrics"]:
        registered = sector_metrics.get(metric["sector"], {})
        expected_unit = registered.get(metric["metric_id"])
        if expected_unit is None:
            raise ValueError(
                f"unregistered company metric {metric['sector']}.{metric['metric_id']}"
            )
        if metric["unit"] != expected_unit:
            raise ValueError(
                f"company metric {metric['metric_id']} uses {metric['unit']}, "
                f"expected {expected_unit}"
            )
        missing_sources = set(metric.get("source_ids", [])) - source_ids
        if missing_sources or not metric.get("source_ids"):
            raise ValueError(
                f"company metric {metric['metric_id']} has invalid sources {sorted(missing_sources)}"
            )
        coverage = metric.get("coverage", {})
        mapped = coverage.get("mapped_activity")
        reported = coverage.get("reported_activity")
        if (
            not isinstance(mapped, (int, float))
            or not isinstance(reported, (int, float))
            or mapped < 0
            or reported <= 0
            or mapped > reported
        ):
            raise ValueError(f"company metric {metric['metric_id']} has invalid coverage")
        if not coverage.get("activity_unit"):
            raise ValueError(f"company metric {metric['metric_id']} lacks a coverage unit")

    for benchmark in alignment["benchmarks"]:
        if benchmark["sector"] not in sector_metrics:
            raise ValueError(
                f"benchmark {benchmark['benchmark_id']} uses an unregistered sector"
            )
        missing_sources = set(benchmark.get("source_ids", [])) - source_ids
        if missing_sources or not benchmark.get("source_ids"):
            raise ValueError(
                f"benchmark {benchmark['benchmark_id']} has invalid sources "
                f"{sorted(missing_sources)}"
            )
        if benchmark["comparison_mode"] == "direct":
            expected_unit = sector_direct_metrics[benchmark["sector"]].get(
                benchmark["metric_id"]
            )
            if expected_unit is None:
                raise ValueError(
                    f"benchmark {benchmark['benchmark_id']} is not a registered direct metric"
                )
            if benchmark.get("unit") != expected_unit:
                raise ValueError(
                    f"benchmark {benchmark['benchmark_id']} uses an incompatible unit"
                )
            if benchmark.get("value") is None or benchmark.get("target_year") is None:
                raise ValueError(
                    f"benchmark {benchmark['benchmark_id']} lacks a direct target value/year"
                )
            if benchmark.get("relation") not in {"at_least", "at_most"}:
                raise ValueError(
                    f"benchmark {benchmark['benchmark_id']} lacks a direct target relation"
                )
        elif benchmark["comparison_mode"] == "contextual":
            if benchmark.get("relation") != "context_only":
                raise ValueError(
                    f"contextual benchmark {benchmark['benchmark_id']} must be context_only"
                )
            value = benchmark.get("value")
            value_min = benchmark.get("value_min")
            value_max = benchmark.get("value_max")
            if (value_min is None) != (value_max is None):
                raise ValueError(
                    f"contextual benchmark {benchmark['benchmark_id']} has a partial range"
                )
            numeric_values = [item for item in (value, value_min, value_max) if item is not None]
            if any(not isinstance(item, (int, float)) or not isfinite(item) for item in numeric_values):
                raise ValueError(
                    f"contextual benchmark {benchmark['benchmark_id']} has a non-finite value"
                )
            if value_min is not None and value_min > value_max:
                raise ValueError(
                    f"contextual benchmark {benchmark['benchmark_id']} has an inverted range"
                )
            if numeric_values and (
                not benchmark.get("unit") or benchmark.get("target_year") is None
            ):
                raise ValueError(
                    f"contextual benchmark {benchmark['benchmark_id']} lacks a unit/year"
                )
            if not numeric_values and not benchmark.get("notes"):
                raise ValueError(
                    f"contextual benchmark {benchmark['benchmark_id']} lacks context"
                )
        else:
            raise ValueError(
                f"benchmark {benchmark['benchmark_id']} has an unknown comparison mode"
            )


def build_contract(payloads: dict[str, dict]) -> dict:
    """Publish the emitted key sets so the web layer can pin the data contract.

    to_json_dict (report/outputs.py) is the single source of truth for the report
    shape; this records it as data, and rejects payloads that disagree on shape.
    """
    reference_slug, reference = next(iter(payloads.items()))
    contract = {
        "firm_result": sorted(reference),
        "cohort": sorted(next(iter(reference["cohorts"].values()))),
        "crossover": sorted({key for row in reference["crossover"] for key in row}),
        "data_quality": sorted(reference["data_quality"]),
        "provenance": sorted(reference["provenance"]),
    }
    for slug, payload in payloads.items():
        if sorted(payload) != contract["firm_result"]:
            raise ValueError(f"contract drift: {slug} top-level keys differ from {reference_slug}")
        for scenario, cohort in payload["cohorts"].items():
            if sorted(cohort) != contract["cohort"]:
                raise ValueError(f"contract drift: {slug}.cohorts.{scenario} keys differ")
        if sorted(payload["data_quality"]) != contract["data_quality"]:
            raise ValueError(f"contract drift: {slug}.data_quality keys differ")
    return contract


def run_firm(_slug: str, fixture_path: Path) -> dict:
    fx = load_fixture(fixture_path)
    result = run(
        fx.firm,
        fx.cohort_year,
        fx.placements,
        fx.countries,
        fx.support,
        fx.config,
        analysis_level=fx.analysis_level,
        layer1_method=fx.layer1_method,
    )
    payload = to_json_dict(result)
    payload["sensitivity"] = run_sensitivity(
        fx.firm, fx.cohort_year, fx.placements, fx.countries, fx.support, fx.config
    )
    payload["inputs"] = effective_inputs(fx, fixture_path)
    return payload


def build_contract_schema() -> dict:
    """Derive the web contract from the internal validation case without publishing it."""
    payload = run_firm("validation", VALIDATION_FIXTURE)
    payload["provenance"] = {
        "engine_version": "",
        "engine_source_sha256": "",
        "workbook": "",
        "workbook_sha256": "",
        "input_sha256": "",
        "dataset_sha256": "",
    }
    return build_contract({"validation": payload})


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    meta = build_meta()
    firms = build_universe()
    countries, support_contract = build_countries()
    alignment = build_alignment_data()
    lifetime = run_lifetime()
    available_company_ids = {row["company_id"] for row in alignment["company_metrics"]}
    cohort_company_ids = {row["company_id"] for row in alignment["product_cohorts"]}
    ready_cohort_ids = {
        row["cohort_id"]
        for row in alignment["impact_readiness"]
        if row["status"] == "available"
    }
    lifetime_company_ids = {
        row["company_id"]
        for row in alignment["product_cohorts"]
        if row["cohort_id"] in ready_cohort_ids
    }
    for firm in firms:
        firm["alignment_available"] = firm["slug"] in available_company_ids
        firm["cohort_available"] = firm["slug"] in cohort_company_ids
        firm["lifetime_result_available"] = firm["slug"] in lifetime_company_ids
        if firm["slug"] == "toyota":
            firm["note"] = TOYOTA_ALIGNMENT_NOTE
        elif firm["slug"] == "hyundai":
            firm["note"] = HYUNDAI_ALIGNMENT_NOTE
        elif firm["slug"] == "jera":
            firm["note"] = JERA_ALIGNMENT_NOTE
        elif firm["slug"] == "koen":
            firm["note"] = KOEN_ALIGNMENT_NOTE
        elif firm["slug"] == "mitsui":
            firm["note"] = MOL_ALIGNMENT_NOTE
    meta["support_contract"] = support_contract
    meta["alignment_contract"] = {
        "version": "alignment-v2",
        "sectors_registered": len(alignment["sectors"]),
        "direct_benchmarks": sum(
            row["comparison_mode"] == "direct" for row in alignment["benchmarks"]
        ),
        "contextual_benchmarks": sum(
            row["comparison_mode"] == "contextual" for row in alignment["benchmarks"]
        ),
        "company_metrics": len(alignment["company_metrics"]),
        "rule": "direct arithmetic requires matching sector, metric definition, geography, and unit",
    }
    meta["impact_contract"] = {
        "version": "export-impact-v1",
        "product_cohorts": len(alignment["product_cohorts"]),
        "cohort_records": sum(
            len(cohort["records"]) for cohort in alignment["product_cohorts"]
        ),
        "destination_pathways": len(alignment["pathways"]),
        "published_lifetime_results": sum(
            row["status"] == "available" for row in alignment["impact_readiness"]
        ),
        "rule": (
            "company × cohort year × destination × product type is computed only after "
            "destination-use, survival, energy, and policy-pathway inputs are sourced"
        ),
    }
    meta["dataset_sha256"] = _json_sha256(
        {
            "countries": countries,
            "firms": firms,
            **alignment,
            "lifetime_results": lifetime,
            "inputs": {},
            "engine_source_sha256": meta["engine_source_sha256"],
            "workbook_sha256": meta["workbook_sha256"],
        }
    )
    expected_names = {
        "contract.json",
        "countries.json",
        "firms.json",
        "meta.json",
        "sectors.json",
        "benchmarks.json",
        "company_metrics.json",
        "product_cohorts.json",
        "pathways.json",
        "impact_readiness.json",
        "sources.json",
        "destination_inputs.json",
        "lifetime_results.json",
    }
    removed = _prune_stale_json_outputs(expected_names)
    if removed:
        print(f"removed stale published outputs: {', '.join(removed)}")
    _write_json(OUT / "contract.json", build_contract_schema())
    _write_json(OUT / "countries.json", countries)
    _write_json(OUT / "firms.json", firms)
    for name, rows in alignment.items():
        _write_json(OUT / f"{name}.json", rows)
    _write_json(OUT / "lifetime_results.json", lifetime)
    _write_json(OUT / "meta.json", meta)
    cohort_count = len(alignment["product_cohorts"])
    lifetime_result_count = sum(
        1 for row in alignment["impact_readiness"] if row["status"] == "available"
    )
    print(
        f"built firms.json ({len(firms)} firms, "
        f"{cohort_count} observed product cohort; "
        f"{lifetime_result_count} lifetime impact result; "
        f"{len(available_company_ids)} companies with supporting evidence) "
        f"and meta.json -> {OUT}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
